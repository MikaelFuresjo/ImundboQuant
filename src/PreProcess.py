"""
MIT License

Copyright (c) [2016] [Mikael Furesjö]
Software = Python Scripts in the [Imundbo Quant v1.6] series

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

IMUNDBO QUANT v1.8 (Preprocessing script)
"""

import csv
from datetime import datetime
import glob
import importlib
import numpy as np
import pandas as pd
import time
import datetime
import matplotlib.pyplot as plt
import os
import pathlib
import sys
import traceback

from config.IQConfig import IQConfig

from gui.console import Console
from utils.CustomTypes import TColumnName, TFeatureLambda, TFeatureLambdasDict
from utils.Utils import columnName, applyFeatures


c = Console(
"""  ___
 |  _ \ _ __ ___ _ __  _ __ ___   ___ ___  ___ ___                         
 | |_) | '__/ _ \ '_ \| '__/ _ \ / __/ _ \/ __/ __|                        
 |  __/| | |  __/ |_) | | | (_) | (_|  __/\__ \__ \                        
 |_|   |_|  \___| .__/|_|  \___/ \___\___||___/___/                        
                |_|                                  

 Preprocess ticker data and extract features to use optimizing machine learning system

""")



# Settings are kept in config/config.json

config = IQConfig()

instrumentsGlob = os.path.join(config.root, config.input.instrumentsGlob)


instruments = glob.glob(instrumentsGlob)
numInstruments = len(instruments)

if numInstruments == 0:
    raise AssertionError("No instruments found in {}".format(instrumentsGlob))

print ('Reading {0} files with ticker data from {1}...'.format(numInstruments, instrumentsGlob))



fileColumns = config.input.columns
supportedColumns = [ "Date", "Open", "High", "Low", "Close", "Volume"]

print("Columns specified: {0}".format(fileColumns))

if not all(column in supportedColumns for column in fileColumns):
    raise AssertionError("Columns not supported specified in config.json/input/columns. \nSpecified: {} \nSupported: {}".format(fileColumns, supportedColumns))


featuresStrategyName = config.features.featuresStrategy
targetsStrategyName = config.features.targetsStrategy

featuresToExtract: TFeatureLambdasDict = []
targetsToExtract: TFeatureLambdasDict = []

try:
    featureStrategy = importlib.import_module("strategies.{}".format(featuresStrategyName))
    print("\nFeatures strategy: {}".format(featuresStrategyName))
    featuresToExtract = featureStrategy.getStrategy()
except Exception as e:
    print("ERROR loading features strategy")
    traceback.print_exc()
    

try:
    targetsStrategy = importlib.import_module("targetStrategies.{}".format(targetsStrategyName))
    print("Targets strategy: {}".format(targetsStrategyName))
    targetsToExtract = targetsStrategy.getStrategy()
except Exception as e:
    print("ERROR loading targets strategy")
    traceback.print_exc()



featuresOutputPath = os.path.join(config.root, config.features.featuresOutputFormat.format(featuresStrategyName))
targetsOutputPath = os.path.join(config.root, config.features.targetsOutputFormat.format(targetsStrategyName))




numCompleted = 0
numFailed = 0


featureColumnNames = []
for feature, settings in featuresToExtract.items():
    featureColumnNames.append(feature)

targetColumnNames = []
for feature, settings in targetsToExtract.items():
    targetColumnNames.append(feature)


filteredFeatureColumnNames = [feature for feature in featureColumnNames if not "Raw" in feature and not "raw" in feature]

print("\nFeatures to calculate: {0}".format(filteredFeatureColumnNames))
print("\nTargets to calculate: {0}".format(targetColumnNames))

input("\nPress [Return] to continue")



allInstrumentsFeatures = None
allInstrumentsTargets = None


for index, instrument in enumerate(instruments):
    instrumentName = os.path.splitext(instrument)[0]

    print("\nProcessing {0} ({1} / {2})...".format(instrumentName, index+1, numInstruments))

    try:
        data = pd.read_csv(instrument, parse_dates=['Date'], header=None, index_col="Date", names=fileColumns, usecols=fileColumns)
        print("Read {0} rows".format(len(data)))

        features = pd.DataFrame(columns = featureColumnNames, index = data.index)
        targets = pd.DataFrame(columns = targetColumnNames, index = data.index)
        numRows = len(data)

    except Exception as e:
        numFailed+=1
        print("Error processing {0}. Skipping".format(instrument))
        traceback.print_exc()
        continue

    applyFeatures(featuresToExtract, data, features)
    applyFeatures(targetsToExtract, data, targets)


    features = features[filteredFeatureColumnNames] #remove any extra
    targets = targets[targetColumnNames] #remove any extra
    
    allInstrumentsFeatures = pd.concat([allInstrumentsFeatures, features], ignore_index = True)
    allInstrumentsTargets = pd.concat([allInstrumentsTargets, targets], ignore_index = True)

    c.timer.print_elapsed("Completed processing of {0}".format(instrumentName))

    numCompleted+=1



        
print("Dropping rows containing, Inf, -Inf, and NaN...")

allInstrumentsFeatures.replace([np.inf, -np.inf], np.nan, inplace = True)
allInstrumentsFeatures.dropna(inplace = True)

allInstrumentsTargets.replace([np.inf, -np.inf], np.nan, inplace = True)
allInstrumentsTargets.dropna(inplace = True)

# intercect indices so that rows removed above are removed from both features and targets
ix = allInstrumentsFeatures.index.intersection(allInstrumentsTargets.index)
allInstrumentsFeatures = allInstrumentsFeatures.loc[ix]
allInstrumentsTargets = allInstrumentsTargets.loc[ix]


print("Saving {0} rows to {1}...".format(len(allInstrumentsFeatures), featuresOutputPath))

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows



pathlib.Path(os.path.dirname(featuresOutputPath + ".msg")).mkdir(parents=True, exist_ok=True) 

allInstrumentsFeatures.to_msgpack(featuresOutputPath + ".msg")
c.timer.print_elapsed("Saved msgpack")
allInstrumentsTargets.to_msgpack(targetsOutputPath + ".msg")
c.timer.print_elapsed("Saved targets msgpack")

allInstrumentsFeatures.to_csv(featuresOutputPath + ".csv", float_format="%.6f")
c.timer.print_elapsed("Saved csv")
allInstrumentsTargets.to_csv(targetsOutputPath + ".csv", float_format="%.6f")
c.timer.print_elapsed("Saved targets csv")

#### Quicker excel export (but still a lot slower than csv and A LOT slower than csv)
#wb = Workbook(write_only=True)
#ws = wb.create_sheet()
#
#for r in dataframe_to_rows(allInstrumentsFeatures, index=False, header=True):
#    ws.append(r)
#wb.save(featuresOutputPath) 
#
#c.timer.print_elapsed("Saved quick excel")
####

### Slower
#allInstrumentsFeatures.to_excel(featuresOutputPath, float_format="%.6f", index=False, freeze_panes=(1,0), sheet_name='report', engine='xlsxwriter') #xlsxwriter

#wb = Workbook()
#wb.new_sheet("Strategy features", data=[allInstrumentsFeatures.columns.tolist(), ] + allInstrumentsFeatures.values.tolist())
#wb.save(featuresOutputPath)

print("Saved {0}".format(featuresOutputPath))
c.timer.print_elapsed('Completed preprocessing {0} files with ticker data ({1} failed) from {2}'.format(numCompleted, numFailed, instrumentsGlob))
